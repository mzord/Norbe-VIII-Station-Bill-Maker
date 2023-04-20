import PySimpleGUI as sg
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from Fetcher import FunctionFetcher


class StationBill:

    # Lista de possíveis nomes para cada função.

    def __init__(self):
        self.workbook = load_workbook("StationBill.xlsm")
        self.stationBill = self.workbook["Lógica"]
        self.alerts = []
        self.craneass_names = ["ASSISTENTE DE GUINDASTEIRO", "ASSISTENTE DE GUINDASTE", "ASSISTENTE OPER DE GUINDASTE", "ASSISTENTE OPER GUINDASTE"]
        self.subseasup_names = ["SUPERVISOR DE SUBSEA", "SUPERVISOR DE SUBSEA SR", "SUBSEA SUPERVISOR"]
        self.toolpusher_names = ["ENCARREGADO DE PLATAFORMA", "ENCARREGADO DE PLATAFORMA (TP)", "ENCARREGADO DE PLATAFORMA (MPD)"]
        self.rstc_names = ["COORD SEGURANÇA TRAB E TREIN"]
        self.captain_names = ["COMANDANTE/OIM", "OIM/COMANDANTE", "OIM COMANDANTE", "COMANDANTE OIM"]
        self.choff_names = ["IMEDIATO", "CHIEF OFFICER"]
        self.driller_names = ["SONDADOR CYBER"]
        self.dpos_names = ["OPERADOR DE POSIC DINAMICO", "OPERADORA DE POSIC DINAMICO"]
        self.ost_names = ["TÉCNICO DE SEGURANÇA"]
        self.drillerass_names = ["ASSISTENTE SONDADOR", "ASSISTENTE DE SONDADOR"]
        self.electricboss_names = ["CHEFE DE ELÉTRICA", "CHEFE DE ELETRICA"]
        self.sdpos_names = ["OPERADOR DE POSIC DINAMICO SR", "OPERADORA DE POSIC DINAMICO SR"]
        self.mnc_names = ["MARINHEIRO DE CONVÉS", "MARINHEIRO DE CONVES"]
        self.cook_names = ["COZINHEIRO (A)", "COZINHEIRO", "COZINHEIRA"]
        self.subseass_names = ["ASSISTENTE DE SUBSEA", "ASSISTENTE SUBSEA"]
        self.derrickmanass_names = ["ASSISTENTE DE TORRISTA"]
        self.rigsup_names = ["SUPERINTENDENTE DE PLATAFORMA"]
        self.maintcoord_names = ["COORD MANUTENÇÃO", "COORDENADOR DE MANUTENÇÃO", "MAINTENANCE COORDINATOR", "COORD DE MANUTENÇÃO"]
        self.elecsup_names = ["SUPERVISOR DE ELETROELETRONICA", "ELECTRICAL SUPERVISOR"]
        self.mechanic_names = ["MECÂNICO", "MECANICO"]
        self.rop_names = ["RADIO OPERADOR", "RADIO OPERADORA"]
        self.craneop_names = ["OPERADOR DE GUINDASTE", "OP. DE GUINDASTE"]
        self.eletronic_names = ["TECNICO DE ELETRONICA", "ELETRONICO"]
        self.oqm_names = ["SEGUNDO OFICIAL DE MÁQUINAS", "SEGUNDO OFICIAL DE MAQUINAS"]
        self.cheng_names = ["CHEFE DE MÁQUINAS", "CHEFE DE MAQUINAS", "CHIEF ENGINEER"]
        self.auxplat_names = ["AUXILIAR DE PLATAFORMA", "AUX. PLATAFORMA"]
        self.deckpusher_names = ["DECKPUSHER", "ENCARREGADO DE CONVÉS", "ENCARREGADO DE CONVES", "DECK PUSHER"]
        self.mnm_names = ["MARINHEIRO DE MÁQUINAS", "MARINHEIRO DE MAQUINAS", "OILER"]
        self.floorman_names = ["PLATAFORMISTA", "FLOORMAN"]
        self.derrickman_names = ["TORRISTA", "DERRICK MAN", "DERRICKMAN"]
        self.campboss_names = ["CAMPBOSS", "NUTRICIONISTA", "COMISSÁRIA", "COMISSÁRIO", "COMISSÁRIO (A)"]
        self.cabin_boy = ["ARRUMADOR", "ARRUMADORA", "ARRUMADOR (A)", "ARRUMADOR(A)"]
        self.mechsup_names = ["SUPERVISOR DE MECÂNICA", "SUPERVISOR DE MECANICA", "MECHANICAL SUPERVISOR", "MECHANICAL SUP", "MECH. SUPERVISOR"]
        self.bosun_names = ["MESTRE DE CABOTAGEM", "BOSUN"]
        self.electricaltech_names = ["ELETRICISTA", "TECNICO DE ELETRICA OFF", "TECNICO DE ELETRICA"]
        self.nurse_names = ["ENFERMEIRO", "ENFERMEIRA", "ENFERMEIRO (A)"]
        sg.theme('DarkAmber')
        self.layout = [
            [sg.Text("Escolha POB desejado:")],
            [sg.Input(), sg.FileBrowse(key="-IN-")],
            [sg.Button("Criar Tabela Mestra"), sg.Button("Cancel")]
        ]
        self.window = sg.Window("Station Bill Maker", self.layout)
        self.make_window()
        

    def set_functions(self, workbook_path):
        self.ff = FunctionFetcher(workbook_path)
        redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')


    # set_captain
        try:
            self.stationBill["E1"].value = self.ff.get_duty(self.captain_names)[0]
        except:
            self.stationBill["E1"].fill = redFill

    # set_choff
        try:
            self.stationBill["H1"].value = self.ff.get_duty(self.choff_names)[0]
        except:
            self.stationBill["H1"].fill = redFill

    # set_dpos
        try:
            self.stationBill["H7"].value = self.ff.get_duty(self.dpos_names)[0]
        except:
            self.stationBill["H7"].fill = redFill
        try:
            self.stationBill["H8"].value = self.ff.get_duty(self.dpos_names)[1]     
        except:
            self.stationBill["H8"].fill = redFill

    # set_sdpos
        try:
            self.stationBill["K7"].value = self.ff.get_duty(self.sdpos_names)[0]
        except:
            self.stationBill["K7"].fill = redFill
        
        try:
            self.stationBill["K8"].value = self.ff.get_duty(self.sdpos_names)[1]
        except:
            self.stationBill["K8"].fill = redFill

    # set_rigsup
        try:
            self.stationBill["E2"].value = self.ff.get_duty(self.rigsup_names)[0]
        except:
            self.stationBill["E2"].fill = redFill

    # set_maintcoord
        try:
            self.stationBill["E3"].value = self.ff.get_duty(self.maintcoord_names)[0]
        except:
            self.stationBill["E3"].fill = redFill

    # set_elecsup
        try:
            self.stationBill["E4"].value = self.ff.get_duty(self.elecsup_names)[0]
        except:
            self.stationBill["E4"].fill = redFill

    # set_rops
        try:
            self.stationBill["N7"].value = self.ff.get_duty(self.rop_names)[0]
        except:
            self.stationBill["N7"].fill = redFill
        try:
            self.stationBill["N8"].value = self.ff.get_duty(self.rop_names)[1]
        except:
            self.stationBill["N8"].fill = redFill

    # set_craneops
        try:
            self.stationBill["E14"].value = self.ff.get_duty(self.craneop_names)[0]
        except:
            self.stationBill["E14"].fill = redFill

        try:
            self.stationBill["E15"].value = self.ff.get_duty(self.craneop_names)[1]
        except:
            self.stationBill["E15"].fill = redFill
        
    # set_mncs
        try:
            self.stationBill["K14"].value = self.ff.get_duty(self.mnc_names)[0]
        except:
            self.stationBill["K14"].fill = redFill
        try:
            self.stationBill["K15"].value = self.ff.get_duty(self.mnc_names)[1]
        except:
            self.stationBill["K15"].fill = redFill

    # set_oqms
        try:
            self.stationBill["E7"].value = self.ff.get_duty(self.oqm_names)[0]
        except:
            self.stationBill["E7"].fill = redFill
        try:
            self.stationBill["E8"].value = self.ff.get_duty(self.oqm_names)[1]
        except:
            self.stationBill["E8"].fill = redFill

    # set_cheng
        try:
            self.stationBill["H2"].value = self.ff.get_duty(self.cheng_names)[0]
        except:
            self.stationBill["H2"].fill = redFill

    # set_auxplat
        try:
            self.stationBill["H50"].value = self.ff.get_duty(self.auxplat_names)[0]
        except:
            self.stationBill["H50"].fill = redFill
        try:
            self.stationBill["H51"].value = self.ff.get_duty(self.auxplat_names)[1]
        except:
            self.stationBill["H51"].fill = redFill
        try:
            self.stationBill["H52"].value = self.ff.get_duty(self.auxplat_names)[2]
        except:
            self.stationBill["H52"].fill = redFill
        try:
            self.stationBill["H53"].value = self.ff.get_duty(self.auxplat_names)[3]
        except:
            self.stationBill["H53"].fill = redFill
        try:
            self.stationBill["H54"].value = self.ff.get_duty(self.auxplat_names)[4]
        except:
            self.stationBill["H54"].fill = redFill
        try:
            self.stationBill["H55"].value = self.ff.get_duty(self.auxplat_names)[5]
        except:
            self.stationBill["H55"].fill = redFill
        try:    
            self.stationBill["H56"].value = self.ff.get_duty(self.auxplat_names)[6]
        except:
            self.stationBill["H56"].fill = redFill
        try:
            self.stationBill["H57"].value = self.ff.get_duty(self.auxplat_names)[7]
        except:
            self.stationBill["H57"].fill = redFill

        #Encarregados de Convés
        try:
            self.stationBill["E21"].value = self.ff.get_duty(self.deckpusher_names)[0]
        except:
            self.stationBill["E21"].fill = redFill
        try:
            self.stationBill["E22"].value = self.ff.get_duty(self.deckpusher_names)[1]
        except:
            self.stationBill["E22"].fill = redFill

        # MNM`s
        try:
            self.stationBill["E28"].value = self.ff.get_duty(self.mnm_names)[0]
        except:
            self.stationBill["E28"].fill = redFill
        try:
            self.stationBill["E29"].value = self.ff.get_duty(self.mnm_names)[1]
        except:
            self.stationBill["E29"].fill = redFill

        # Torristas
        try:
            self.stationBill["E35"].value = self.ff.get_duty(self.derrickman_names)[0]
        except:
            self.stationBill["E35"].fill = redFill
        try:
            self.stationBill["E36"].value = self.ff.get_duty(self.derrickman_names)[1]
        except:
            self.stationBill["E36"].fill = redFill

        # Nutricionista
        try:
            self.stationBill["E42"].value = self.ff.get_duty(self.campboss_names)[0]
        except:
            self.stationBill["E42"].fill = redFill

        # Arrumador (a)
        try:
            self.stationBill["E50"].value = self.ff.get_duty(self.cabin_boy)[0]
        except:
            self.stationBill["E50"].fill = redFill
        try:
            self.stationBill["E51"].value = self.ff.get_duty(self.cabin_boy)[1]
        except:
            self.stationBill["E51"].fill = redFill
        try:
            self.stationBill["E52"].value = self.ff.get_duty(self.cabin_boy)[2]
        except:
            self.stationBill["E52"].fill = redFill
        try:
            self.stationBill["E53"].value = self.ff.get_duty(self.cabin_boy)[3]
        except:
            self.stationBill["E53"].fill = redFill
        try:
            self.stationBill["E54"].value = self.ff.get_duty(self.cabin_boy)[4]
        except:
            self.stationBill["E54"].fill = redFill
        try:
            self.stationBill["E55"].value = self.ff.get_duty(self.cabin_boy)[5]
        except:
            self.stationBill["E55"].fill = redFill

        # Supervisor de Mecânica
        try:
            self.stationBill["H4"].value = self.ff.get_duty(self.mechsup_names)[0]
        except:
            self.stationBill["H4"].fill = redFill

        # Supervisor de Subsea
        try:
            self.stationBill["N1"].value = self.ff.get_duty(self.subseasup_names)[0]
        except:
            self.stationBill["N1"].fill = redFill
        try:
            self.stationBill["N2"].value = self.ff.get_duty(self.subseasup_names)[1]
        except:
            self.stationBill["N2"].fill = redFill
        
        # Assistente de Guindasteiro
        try:
            self.stationBill["H14"].value = self.ff.get_duty(self.craneass_names)[0]
        except:
            self.stationBill["H14"].fill = redFill
        try:
            self.stationBill["H15"].value = self.ff.get_duty(self.craneass_names)[1]
        except:
            self.stationBill["H15"].fill = redFill

        # Mestre de Cabotagem (Precisa de 1)
        try:
            self.stationBill["N14"].value = self.ff.get_duty(self.bosun_names)[0]
        except:
            self.stationBill["N14"].fill = redFill
        try:
            self.stationBill["N15"].value = self.ff.get_duty(self.bosun_names)[1]
        except:
            self.stationBill["N15"].fill = redFill

        # Eletricista (Precisa de 2)
        try:
            self.stationBill["H21"].value = self.ff.get_duty(self.electricaltech_names)[0]
        except:
            self.stationBill["H21"].fill = redFill
        try:
            self.stationBill["H22"].value = self.ff.get_duty(self.electricaltech_names)[1]
        except:
            self.stationBill["H22"].fill = redFill

        # Eletronico (Precisa de 1)
        try:
            self.stationBill["K21"].value = self.ff.get_duty(self.eletronic_names)[0]
        except:
            self.stationBill["K21"].fill = redFill
        try:
            self.stationBill["K22"].value = self.ff.get_duty(self.eletronic_names)[1]
        except:
            self.stationBill["K22"].fill = redFill

        # Mecanico (Precisa de 2)
        try:
            self.stationBill["N21"].value = self.ff.get_duty(self.mechanic_names)[0]
        except:
            self.stationBill["N21"].fill = redFill
        try:
            self.stationBill["N22"].value = self.ff.get_duty(self.mechanic_names)[1]
        except:
            self.stationBill["N22"].fill = redFill

        # Encarregado de Plataforma (Precisa de 2)
        try:
            self.stationBill["H28"].value = self.ff.get_duty(self.toolpusher_names)[0]
        except:
            self.stationBill["H28"].fill = redFill
        try:
            self.stationBill["H29"].value = self.ff.get_duty(self.toolpusher_names)[1]
        except:
            self.stationBill["H29"].fill = redFill

        # Sondador Cyber (Precisa de 2)
        try:
            self.stationBill["K28"].value = self.ff.get_duty(self.driller_names)[0]
        except:
            self.stationBill["K28"].fill = redFill
        try:
            self.stationBill["K29"].value = self.ff.get_duty(self.driller_names)[1]
        except:
            self.stationBill["K29"].fill = redFill

        # Assistente de Sondador (Precisa de 2)
        try:
            self.stationBill["N28"].value = self.ff.get_duty(self.drillerass_names)[0]
        except:
            self.stationBill["N28"].fill = redFill
        try:
            self.stationBill["N29"].value = self.ff.get_duty(self.drillerass_names)[1]
        except:
            self.stationBill["N29"].fill = redFill

        # Assistente de Torrista (Precisa de 2)
        try:
            self.stationBill["H35"].value = self.ff.get_duty(self.derrickmanass_names)[0]
        except:
            self.stationBill["H35"].fill = redFill
        try:
            self.stationBill["H36"].value = self.ff.get_duty(self.derrickmanass_names)[1]
        except:
            self.stationBill["H36"].fill = redFill

        # Plataformista (Precisa de 3)
        try:
            self.stationBill["K35"].value = self.ff.get_duty(self.floorman_names)[0]
        except:
            self.stationBill["K35"].fill = redFill
        try:
            self.stationBill["K36"].value = self.ff.get_duty(self.floorman_names)[1]
        except:
            self.stationBill["K36"].fill = redFill
        try:
            self.stationBill["K37"].value = self.ff.get_duty(self.floorman_names)[2]
        except:
            self.stationBill["K37"].fill = redFill
        try:
            self.stationBill["K38"].value = self.ff.get_duty(self.floorman_names)[3]
        except:
            self.stationBill["K38"].fill = redFill

        # Assistente de Subsea (Precisa de 1)
        try:
            self.stationBill["N35"].value = self.ff.get_duty(self.subseass_names)[0]
        except:
            self.stationBill["N35"].fill = redFill

        # Cozinheiro (Precisa de 1)
        try:
            self.stationBill["H42"].value = self.ff.get_duty(self.cook_names)[0]
        except:
            self.stationBill["H42"].fill = redFill

        # Técnico de Segurança (Precisa de 2)
        try:
            self.stationBill["N42"].value = self.ff.get_duty(self.ost_names)[0]
        except:
            self.stationBill["N42"].fill = redFill
        try:
            self.stationBill["N43"].value = self.ff.get_duty(self.ost_names)[1]
        except:
            self.stationBill["N43"].fill = redFill

        # Enfermeiro
        try:
            self.stationBill["K50"].value = self.ff.get_duty(self.nurse_names)[0]
        except:
            self.stationBill["K50"].fill = redFill

        # Chefe de Elétrica (Precisa de 1)
        try:
            self.stationBill["H3"].value = self.ff.get_duty(self.electricboss_names)[0]
        except:
            self.stationBill["H3"].fill = redFill

        # RSTC (Preicsa de 1)
        try:
            self.stationBill["K2"].value = self.ff.get_duty(self.rstc_names)[0]
        except:
            self.stationBill["K2"].fill = redFill

    # Método em progresso
    def check_errors(self):
        if len(self.ff.get_duty(self.ost_names)) >2:
            print("\nMais de dois OST`s.")
            print(self.ff.get_duty(self.ost_names))

        if len(self.ff.get_duty(self.nurse_names)) > 1:
            print("\nMais de um enfermeiro a bordo?")
            print(self.ff.get_duty(self.nurse_names))

        if len(self.ff.get_duty(self.cook_names)) >1:
            print("\nMais de um cozinheiro a bordo?")
            print(self.ff.get_duty(self.cook_names))

        if len(self.ff.get_duty(self.subseass_names)) >1:
            print("\nMais de um assistente de subsea a bordo.")
            print(self.ff.get_duty(self.subseass_names))

        if len(self.ff.get_duty(self.floorman_names)) >3:
            print("\nMais de três plataformistas.")
            print(self.ff.get_duty(self.floorman_names))

    # Função que gera a GUI do PySimpleGUI
    def make_window(self):
        while True:
            event, values = self.window.read()
            if event == sg.WIN_CLOSED or event == "Cancel":
                break
            if event == "Criar Tabela Mestra":
                self.set_functions(values["-IN-"])
                self.check_errors()
                self.save()
                self.window.close()
        self.window.close()
        
    # Função que salva o que foi gerado.
    def save(self):
        self.workbook.save("./StationBillMade.xlsx")


sb = StationBill()