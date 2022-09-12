from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from Fetcher import FunctionFetcher


class StationBill:

    def __init__(self):
        self.workbook = load_workbook("StationBill.xlsm")
        self.stationBill = self.workbook["PORTUGUÊS"]
        self.ff = FunctionFetcher()

        self.captain_names = ["COMANDANTE/OIM", "OIM/COMANDANTE", "OIM COMANDANTE", "COMANDANTE OIM"]
        self.choff_names = ["IMEDIATO", "CHIEF OFFICER"]
        self.dpos_names = ["OPERADOR DE POSIC DINAMICO", "OPERADORA DE POSIC DINAMICO"]
        self.sdpos_names = ["OPERADOR DE POSIC DINAMICO SR", "OPERADORA DE POSIC DINAMICO SR"]
        self.mnc_names = ["MARINHEIRO DE CONVÉS", "MARINHEIRO DE CONVES"]
        self.rigsup_names = ["SUPERINTENDENTE DE PLATAFORMA"]
        self.maintcoord_names = ["COORD MANUTENÇÃO", "COORDENADOR DE MANUTENÇÃO", "MAINTENANCE COORDINATOR", "COORD DE MANUTENÇÃO"]
        self.elecsup_names = ["SUPERVISOR DE ELETROELETRONICA", "ELECTRICAL SUPERVISOR"]
        self.rop_names = ["RADIO OPERADOR", "RADIO OPERADORA"]
        self.craneop_names = ["OPERADOR DE GUINDASTE", "OP. DE GUINDASTE"]
        self.oqm_names = ["SEGUNDO OFICIAL DE MÁQUINAS", "SEGUNDO OFICIAL DE MAQUINAS"]
        self.cheng_names = ["CHEFE DE MÁQUINAS", "CHEFE DE MAQUINAS", "CHIEF ENGINEER"]
        self.auxplat_names = ["AUXILIAR DE PLATAFORMA", "AUX. PLATAFORMA"]

    def set_functions(self):
        
        redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')


    # set_captain
        try:
            self.stationBill["C20"].value = self.ff.get_duty(self.captain_names)[0]
        except:
            self.stationBill["C20"].fill = redFill

    # set_choff
        try:
            self.stationBill["AB27"].value = self.ff.get_duty(self.choff_names)[0]
        except:
            self.stationBill["AB27"].fill = redFill

    # set_dpo1
        try:
            self.stationBill["C24"].value = self.ff.get_duty(self.dpos_names)[0]
        except:
            self.stationBill["C24"].fill = redFill
        try:
            self.stationBill["T19"].value = self.ff.get_duty(self.dpos_names)[0]     
        except:
            self.stationBill["T19"].fill = redFill

    # set_dpo2
        try:        
            self.stationBill["C30"].value = self.ff.get_duty(self.dpos_names)[1]
        except:
            self.stationBill["C30"].fill = redFill
        try:
            self.stationBill["AB28"].value = self.ff.get_duty(self.dpos_names)[1]     
        except:
            self.stationBill["AB28"].fill = redFill

    # set_sdpo1
        try:
            self.stationBill["C23"].value = self.ff.get_duty(self.sdpos_names)[0]
        except:
            self.stationBill["C23"].fill = redFill
        
        try:
            self.stationBill["T18"].value = self.ff.get_duty(self.sdpos_names)[0]
        except:
            self.stationBill["T18"].fill = redFill
    
    # set_sdpo2
        try:
            self.stationBill["C25"].value = self.ff.get_duty(self.sdpos_names)[1]
        except:
            self.stationBill["C25"].fill = redFill
        try:
            self.stationBill["AB18"].value = self.ff.get_duty(self.sdpos_names)[1]
        except:
            self.stationBill["AB18"].fill = redFill

    # set_rigsup
        try:
            self.stationBill["C21"].value = self.ff.get_duty(self.rigsup_names)[0]
        except:
            self.stationBill["C21"].fill = redFill

    # set_maintcoord
        try:
            self.stationBill["C22"].value = self.ff.get_duty(self.maintcoord_names)[0]
        except:
            self.stationBill["C22"].fill = redFill

    # set_elecsup
        try:
            self.stationBill["C27"].value = self.ff.get_duty(self.elecsup_names)[0]
        except:
            self.stationBill["T20"].fill = redFill

    # set_rop1
        try:
            self.stationBill["C28"].value = self.ff.get_duty(self.rop_names)[0]
        except:
            self.stationBill["C28"].fill = redFill

    # set_rop2
        try:
            self.stationBill["C29"].value = self.ff.get_duty(self.rop_names)[1]
        except:
            self.stationBill["C29"].fill = redFill

    # set_craneop1
        try:
            self.stationBill["T20"].value = self.ff.get_duty(self.craneop_names)[0]
        except:
            self.stationBill["T20"].fill = redFill

        try:
            self.stationBill["E38"].value = self.ff.get_duty(self.craneop_names)[0]
        except:
            self.stationBill["E38"].fill = redFill
        try:
            self.stationBill["E66"].value = self.ff.get_duty(self.craneop_names)[0]
        except:
            self.stationBill["E66"].fill = redFill

    # set_craneop2
        try:
            self.stationBill["AB29"].value = self.ff.get_duty(self.craneop_names)[1]
        except:
            self.stationBill["AB29"].fill = redFill

        try:
            self.stationBill["M38"].value = self.ff.get_duty(self.craneop_names)[1]
        except:
            self.stationBill["M38"].fill = redFill

    # set_mnc1
        try:
            self.stationBill["T21"].value = self.ff.get_duty(self.mnc_names)[0]
        except:
            self.stationBill["T21"].fill = redFill
        try:
            self.stationBill["M63"].value = self.ff.get_duty(self.mnc_names)[0]
        except:
            self.stationBill["M63"].fill = redFill
    
    # set_mnc2
        try:
            self.stationBill["AB30"].value = self.ff.get_duty(self.mnc_names)[1]
        except:
            self.stationBill["AB30"].fill = redFill

    # set_oqm1
        try:
            self.stationBill["AB22"].value = self.ff.get_duty(self.oqm_names)[0]
        except:
            self.stationBill["AB22"].fill = redFill
        try:
            self.stationBill["M46"].value = self.ff.get_duty(self.oqm_names)[0]
        except:
            self.stationBill["M46"].fill = redFill

    # set_oqm2
        try:
            self.stationBill["T22"].value = self.ff.get_duty(self.oqm_names)[1]
        except:
            self.stationBill["T22"].fill = redFill

    # set_cheng
        try:
            self.stationBill["AB31"].value = self.ff.get_duty(self.cheng_names)[0]
        except:
            self.stationBill["AB31"].fill = redFill
        try:
            self.stationBill["M45"].value = self.ff.get_duty(self.cheng_names)[0]
        except:
            self.stationBill["M45"].fill = redFill

    # set_auxplat
        try:
            self.stationBill["Q35"].value = self.ff.get_duty(self.auxplat_names)[0]
        except:
            self.stationBill["Q35"].fill = redFill
        try:
            self.stationBill["Y35"].value = self.ff.get_duty(self.auxplat_names)[1]
        except:
            self.stationBill["Y35"].fill = redFill
        try:
            self.stationBill["Q36"].value = self.ff.get_duty(self.auxplat_names)[2]
        except:
            self.stationBill["Q36"].fill = redFill
        try:
            self.stationBill["Y36"].value = self.ff.get_duty(self.auxplat_names)[3]
        except:
            self.stationBill["Y36"].fill = redFill
        try:
            self.stationBill["Q37"].value = self.ff.get_duty(self.auxplat_names)[4]
        except:
            self.stationBill["Q37"].fill = redFill
        try:
            self.stationBill["Y37"].value = self.ff.get_duty(self.auxplat_names)[5]
        except:
            self.stationBill["Y37"].fill = redFill
        try:    
            self.stationBill["Q38"].value = self.ff.get_duty(self.auxplat_names)[6]
        except:
            self.stationBill["Q38"].fill = redFill
        try:
            self.stationBill["Y38"].value = self.ff.get_duty(self.auxplat_names)[7]
        except:
            self.stationBill["Y38"].fill = redFill


    def save(self):
        self.workbook.save("./StationBillMade.xlsx")


sb = StationBill()
sb.set_functions()
sb.save()